import pandas as pd
import openpyxl
import calculation as calc

count = 1
index = range(1, len(calc.all_pressure)+1)
heading = [
    "start node",
    "end node",
    "start node pressure",
    "end node pressure",
    "flow",
    "pipe diameter",
    "pipe length",
    "head loss",
]
headings_data = []
for n in index:
    if n <= calc.no_of_sprinkler_on_branch_lines + 1:
        if n < calc.no_of_sprinkler_on_branch_lines - 2:
            entry = [
                n,
                n + 1,
                calc.all_pressure[n - 1],
                calc.all_pressure[n],
                calc.all_flowrate[n - 1],
                calc.br_line_int_dia,
                calc.DIST_OF_SPRINKLER_ON_BRANCH_LINES,
            ]
            entry.append(entry[3] - entry[2])
            headings_data.append(entry)

        elif n == calc.no_of_sprinkler_on_branch_lines - 2:
            entry = [
                n,
                n + 1,
                calc.all_pressure[n - 1],
                calc.all_pressure[n],
                calc.all_flowrate[n - 1],
                calc.br2_line_int_dia,
                calc.DIST_OF_SPRINKLER_ON_BRANCH_LINES,
            ]
            entry.append(entry[3] - entry[2])
            headings_data.append(entry)

        elif n == calc.no_of_sprinkler_on_branch_lines - 1:
            length = calc.equiv_length_t(
                l=calc.DIST_OF_SPRINKLER_ON_BRANCH_LINES / 2,
                d_1=calc.br2_line_int_dia,
                d_2=calc.int_dia_sch40_2inch,
                fitt_equiv=calc.t_fitt_sch40_2inch,
            )

            entry = [
                n,
                n + 1,
                calc.all_pressure[n - 1],
                calc.all_pressure[n],
                calc.all_flowrate[n - 1],
                calc.br2_line_int_dia,
                length,
            ]
            entry.append(entry[3] - entry[2])
            headings_data.append(entry)

        elif n == calc.no_of_sprinkler_on_branch_lines:
            length = calc.equiv_length_t(
                l=calc.dist_between_branch_lines,
                d_1=calc.cr_line_int_dia,
                d_2=calc.int_dia_sch40_3inch,
                fitt_equiv=calc.t_fitt_sch40_3inch,
            )
            entry = [
                n,
                n + (calc.no_of_sprinkler_on_branch_lines + 1),
                calc.all_pressure[n - 1],
                calc.all_pressure[n + calc.no_of_sprinkler_on_branch_lines],
                calc.all_flowrate[n],
                calc.br_line_int_dia,
                length,
            ]
            entry.append(entry[3] - entry[2])
            headings_data.append(entry)
        elif n == calc.no_of_sprinkler_on_branch_lines + 1:
            length = calc.equiv_length_t(
                l=calc.DIST_OF_SPRINKLER_ON_BRANCH_LINES / 2,
                d_1=calc.br_line_int_dia,
                d_2=calc.int_dia_sch40_1_5inch,
                fitt_equiv=calc.t_fitt_sch40_1_5inch,
            )

            entry = [
                n,
                n - 1,
                calc.all_pressure[n - 1],
                calc.all_pressure[n - 2],
                calc.all_flowrate[n - 2],
                calc.br_line_int_dia,
                length,
            ]
            entry.append(entry[3] - entry[2])
            headings_data.append(entry)
            count += 1

    elif n > calc.no_of_sprinkler_on_branch_lines + 1:
        p = n % (calc.no_of_sprinkler_on_branch_lines + 1)
        if p == 0:
            p = calc.no_of_sprinkler_on_branch_lines + 1
        if p <= calc.no_of_sprinkler_on_branch_lines + 1:
            if p < calc.no_of_sprinkler_on_branch_lines - 2:
                entry = [
                    n,
                    n + 1,
                    calc.all_pressure[n - 1],
                    calc.all_pressure[n],
                    calc.all_flowrate[n - 1],
                    calc.br_line_int_dia,
                    calc.DIST_OF_SPRINKLER_ON_BRANCH_LINES,
                ]
                entry.append(entry[3] - entry[2])
                headings_data.append(entry)

            elif p == calc.no_of_sprinkler_on_branch_lines - 2:
                entry = [
                    n,
                    n + 1,
                    calc.all_pressure[n - 1],
                    calc.all_pressure[n],
                    calc.all_flowrate[n - 1],
                    calc.br2_line_int_dia,
                    calc.DIST_OF_SPRINKLER_ON_BRANCH_LINES,
                ]
                entry.append(entry[3] - entry[2])
                headings_data.append(entry)

            elif p == calc.no_of_sprinkler_on_branch_lines - 1:
                length = calc.equiv_length_t(
                    l=calc.DIST_OF_SPRINKLER_ON_BRANCH_LINES / 2,
                    d_1=calc.br2_line_int_dia,
                    d_2=calc.int_dia_sch40_2inch,
                    fitt_equiv=calc.t_fitt_sch40_2inch,
                )

                entry = [
                    n,
                    n + 1,
                    calc.all_pressure[n - 1],
                    calc.all_pressure[n],
                    calc.all_flowrate[n - 1],
                    calc.br2_line_int_dia,
                    length,
                ]
                entry.append(entry[3] - entry[2])
                headings_data.append(entry)

            elif p == calc.no_of_sprinkler_on_branch_lines:
                print(count)
                # print(n)
                print(f"no of branch {calc.no_of_branch_lines}")
                if count < calc.no_of_branch_lines:
                    length = calc.equiv_length_t(
                        l=calc.dist_between_branch_lines,
                        d_1=calc.cr_line_int_dia,
                        d_2=calc.int_dia_sch40_3inch,
                        fitt_equiv=calc.t_fitt_sch40_3inch,
                    )

                    print(n)
                    print(count)
                    entry = [
                        n,
                        n + (calc.no_of_sprinkler_on_branch_lines + 1),
                        calc.all_pressure[n - 1],
                        calc.all_pressure[n + calc.no_of_sprinkler_on_branch_lines],
                        calc.all_flowrate[n],
                        calc.br_line_int_dia,
                        length,
                    ]
                    entry.append(entry[3] - entry[2])
                headings_data.append(entry)
            elif p == calc.no_of_sprinkler_on_branch_lines + 1:
                length = calc.equiv_length_t(
                    l=calc.DIST_OF_SPRINKLER_ON_BRANCH_LINES / 2,
                    d_1=calc.br_line_int_dia,
                    d_2=calc.int_dia_sch40_1_5inch,
                    fitt_equiv=calc.t_fitt_sch40_1_5inch,
                )

                entry = [
                    n,
                    n - 1,
                    calc.all_pressure[n - 1],
                    calc.all_pressure[n - 2],
                    calc.all_flowrate[n - 2],
                    calc.br_line_int_dia,
                    length,
                ]
                entry.append(entry[3] - entry[2])
                headings_data.append(entry)
                count += 1

# print(headings_data)
print(len(headings_data))


df = pd.DataFrame(headings_data, columns=heading, index=range(1, len(headings_data)+1))

df.to_excel("pipe_data.xlsx", sheet_name="pipe_data")
# index = np.arange(1, len(all_pressure)+1)
# output_dict =
# for n in range(1, len(all_pressure)+1):
#     output_dict = {"start node": {n:}}
#     df.to_csv('output/pre


headings_data2 = []
index2 = range(1, len(calc.all_pressure) + 1)
# print(len(calc.all_pressure))
# print(len(calc.all_discharge_rate))
heading2 = [
    "node point",
    "pressure",
    "discharge rate",
]
for n in index:
    entry = [n, calc.all_pressure[n - 1], calc.all_discharge_rate[n - 1]]

    headings_data2.append(entry)

# print(len(headings_data2))
# print(headings_data2)
df2 = pd.DataFrame(headings_data2, columns=heading2, index=index2)


wb = openpyxl.Workbook()
wb.create_sheet("pipe_data")
wb.create_sheet("node_data")
wb.save("pipe_data.xlsx")

with pd.ExcelWriter("pipe_data.xlsx") as writer:
    df.to_excel(writer, sheet_name="pipe_data", index=0)
    df2.to_excel(writer, sheet_name="node_data", index=1)


