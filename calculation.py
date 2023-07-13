import math
import pandas as pd
import openpyxl


def full_calc(values):
    for i, (key, value) in enumerate(values.items()):
        if key == "Design area selected (ft²):":
            DESIGN_AREA = float(value)
        elif key == "Length of study area (ft):":
            LENGTH = float(value)
        elif key == "Width of study area (ft):":
            WIDTH = float(value)
        elif key == "Sprinkler coefficient:":
            k = float(value)
        elif key == "Occupancy hazard type":
            HAZARD_TYPE = value
        elif key == "Ceiling Elevation:":
            elevation = float(value)
    if HAZARD_TYPE == "Ordinary Hazard (group 1)":
        density = (0.03 * (3000 - DESIGN_AREA) / 1500) + 0.12
    elif HAZARD_TYPE == "Ordinary Hazard (group 2)":
        density = (0.03 * (3000 - DESIGN_AREA) / 1500) + 0.12

    hose_stream_demand = 250
    MAX_COVERAGE = 120
    MAX_SPACING = 12
    # the maximum spacing basd of nfpa standard

    c_pipe = 120  # friction loss coefficient for wet systems

    br_line_int_dia = 1.682
    br2_line_int_dia = 2.157
    cr_line_int_dia = 3.260
    int_dia_sch40_1_5inch = 1.610
    int_dia_sch40_2inch = 2.067
    int_dia_sch40_3inch = 3.068
    t_fitt_sch40_3inch = 15
    t_fitt_sch40_2inch = 10
    t_fitt_sch40_1_5inch = 8

    def x_round(x):
        return round(x * 4) / 4

    no_of_branch_lines = math.ceil(
        WIDTH / MAX_SPACING
    )  # this is the number of branch lines
    dist_between_branch_lines = (
        WIDTH / no_of_branch_lines
    )  # this is the distance between branch lines
    dist_between_branch_lines = x_round(dist_between_branch_lines)
    branch_line_wall_dist = (
        WIDTH - (no_of_branch_lines * dist_between_branch_lines / 2)
    ) / 2

    no_of_sprinkler_on_branch_lines = math.ceil(
        LENGTH / MAX_SPACING
    )  # this is the number of sprinklers on branch lines
    DIST_OF_SPRINKLER_ON_BRANCH_LINES = (
        LENGTH / no_of_sprinkler_on_branch_lines
    )  # this is spacing_between_sprinklers
    dist_between_branch_lines = x_round(dist_between_branch_lines)
    last_sprinkler_wall_dist = (
        LENGTH
        - ((no_of_sprinkler_on_branch_lines - 1) * DIST_OF_SPRINKLER_ON_BRANCH_LINES)
    ) / 2

    calc_coverage_area = dist_between_branch_lines * DIST_OF_SPRINKLER_ON_BRANCH_LINES

    calc_coverage_area_check = True
    calc_design_area_check = True

    while calc_coverage_area_check is True:
        ### this is the while loop to calculate the number of sprinklers in the design area ###
        if calc_coverage_area < MAX_COVERAGE:
            branch_line_wall_dist = (
                WIDTH - (no_of_branch_lines * dist_between_branch_lines / 2)
            ) / 2
            last_sprinkler_wall_dist = (
                LENGTH
                - (
                    (no_of_sprinkler_on_branch_lines - 1)
                    * DIST_OF_SPRINKLER_ON_BRANCH_LINES
                )
            ) / 2
            largest_area = (
                branch_line_wall_dist
                + (dist_between_branch_lines / 2)
                + last_sprinkler_wall_dist
                + (DIST_OF_SPRINKLER_ON_BRANCH_LINES / 2)
            )
            if (
                dist_between_branch_lines * DIST_OF_SPRINKLER_ON_BRANCH_LINES
                > largest_area
            ):
                largest_area = (
                    dist_between_branch_lines * DIST_OF_SPRINKLER_ON_BRANCH_LINES
                )
            if largest_area < MAX_COVERAGE:
                min_length = 1.2 * math.sqrt(DESIGN_AREA)
                min_sprinkler_on_branch = math.ceil(
                    (min_length / DIST_OF_SPRINKLER_ON_BRANCH_LINES)
                )
                actual_length = (
                    min_sprinkler_on_branch - 0.5
                ) * DIST_OF_SPRINKLER_ON_BRANCH_LINES + last_sprinkler_wall_dist
                width = DESIGN_AREA / actual_length
                sprinkler_across_branch = math.ceil((width / dist_between_branch_lines))
                calc_design_area = actual_length * width
                add_node = 0
                while calc_design_area_check is True:
                    if calc_design_area < DESIGN_AREA:
                        calc_design_area += calc_coverage_area
                        add_node += 1
                    else:
                        calc_design_area_check = False
                no_of_sprinkler = (
                    min_sprinkler_on_branch * sprinkler_across_branch
                ) + add_node
            calc_coverage_area_check = False
        else:
            no_of_branch_lines += 1
            dist_between_branch_lines = WIDTH / no_of_branch_lines
            calc_coverage_area = (
                dist_between_branch_lines * DIST_OF_SPRINKLER_ON_BRANCH_LINES
            )  # final computation for the coverage area of a single sprinkler

    all_discharge_rate = []
    flow_rate_end = 0
    all_flowrate = []
    all_pressure = []
    inverted_pressure_add = []
    inverted_discharge_add = []
    inverted_flowrate = []
    pressure_add = 0
    og_flowrate_end = 0
    v = 0

    def pressure(q, k=k):
        """compute for the pressure"""
        pressure = (q / k) ** 2
        return pressure

    def discharge_rate(p, k=k):
        """compute for the discharge rate"""
        discharge_rate = k * math.sqrt(p)
        return discharge_rate

    node1_discharge_rate = calc_coverage_area * density
    node1_pressure = pressure(q=node1_discharge_rate)

    def pressure_drop(q, l_eqv, d, C=c_pipe):
        """compute for the pressure drop using hazen williams equation"""
        p_drop = 4.52 * q**1.85 * l_eqv / (C**1.85 * d**4.87)
        return p_drop

    def equiv_length_t(l, d_1, d_2, fitt_equiv):
        """compute for the equivalent length of the pipe after considering equivalent length of fittings"""
        equiv_length_t = l + ((d_1 / d_2) ** 4.87 * fitt_equiv)
        return equiv_length_t

    def horsepower_calc(q, p):
        """compute for the horsepower"""
        p *= 2.31
        horsepower = (q * p) / (3960 * 0.63)
        return horsepower

    def suitable_pump(number):
        closest_number = None
        min_difference = float("inf")  # Initialize with a large value
        values = [
            25,
            50,
            100,
            150,
            200,
            250,
            300,
            400,
            450,
            500,
            750,
            1000,
            1250,
            1500,
            2000,
            2500,
            3000,
            3500,
            4000,
            4500,
            5000,
        ]
        for value in values:
            difference = abs(number - value)
            if difference < min_difference and number >= value:
                min_difference = difference
                closest_number = value
            if closest_number == None:
                closest_number = 25
        return closest_number
        
    
    for m in range(1, no_of_branch_lines + 1):
        """this is the for loop for the branch lines"""
        if m == 1:
            """this is the if statement for the most remote sprinkler"""
            all_pressure.append(node1_pressure)
            all_discharge_rate.append(node1_discharge_rate)
            pressure = node1_pressure
            flow_rate_end += node1_discharge_rate
            all_flowrate.append(flow_rate_end)
            for n in range(2, no_of_sprinkler_on_branch_lines + 2):
                """this is the for loop for the sprinklers on the most remote branch line"""

                if n < no_of_sprinkler_on_branch_lines:
                    if n < no_of_sprinkler_on_branch_lines - 1:
                        p_drop = pressure_drop(
                            q=flow_rate_end,
                            l_eqv=DIST_OF_SPRINKLER_ON_BRANCH_LINES,
                            d=br_line_int_dia,
                        )
                        pressure += p_drop
                        discharge = discharge_rate(p=pressure)
                        all_discharge_rate.append(discharge)
                        flow_rate_end += discharge
                        all_flowrate.append(flow_rate_end)
                        all_pressure.append(pressure)
                        
                    else:
                        p_drop = pressure_drop(
                            q=flow_rate_end,
                            l_eqv=DIST_OF_SPRINKLER_ON_BRANCH_LINES,
                            d=br2_line_int_dia,
                        )
                        pressure += p_drop
                        discharge = discharge_rate(p=pressure)
                        all_discharge_rate.append(discharge)
                        flow_rate_end += discharge
                        all_flowrate.append(flow_rate_end)
                        all_pressure.append(pressure)

                elif n == no_of_sprinkler_on_branch_lines:
                    length = equiv_length_t(
                        l=DIST_OF_SPRINKLER_ON_BRANCH_LINES / 2,
                        d_1=br2_line_int_dia,
                        d_2=int_dia_sch40_2inch,
                        fitt_equiv=t_fitt_sch40_2inch,
                    )
                    p_drop = pressure_drop(
                        q=flow_rate_end, l_eqv=length, d=br2_line_int_dia
                    )
                    pressure += p_drop
                    all_pressure.append(pressure)
                    all_discharge_rate.append(0)
                    
                elif n == no_of_sprinkler_on_branch_lines + 1:
                    discharge_last = node1_discharge_rate
                    pressure_last = node1_pressure
                    length = equiv_length_t(
                        l=DIST_OF_SPRINKLER_ON_BRANCH_LINES / 2,
                        d_1=br_line_int_dia,
                        d_2=int_dia_sch40_1_5inch,
                        fitt_equiv=t_fitt_sch40_1_5inch,
                    )
                    p_drop = pressure_drop(
                        q=discharge_last, l_eqv=length, d=br_line_int_dia
                    )
                    discharge_last = (
                        math.sqrt(all_pressure[len(all_pressure) - 1])
                        * discharge_last
                        / math.sqrt(pressure_last + p_drop)
                    )
                    all_discharge_rate.append(discharge_last)
                    all_flowrate.append(discharge_last)
                    p_drop = pressure_drop(
                        q=discharge_last, l_eqv=length, d=br_line_int_dia
                    )
                    all_pressure.append(all_pressure[len(all_pressure) - 1] - p_drop)
                    flow_rate_end += discharge_last

                    main_pressure = all_pressure
                    main_discharge_rate = all_discharge_rate

        elif m > 1:
            for n in range(no_of_sprinkler_on_branch_lines, 0, -1):
                """computation for presuure changebetween two sucessive branchline"""
                if n == no_of_sprinkler_on_branch_lines:
                    inverted_discharge_add.append(0)

                    length = equiv_length_t(
                        l=dist_between_branch_lines,
                        d_1=cr_line_int_dia,
                        d_2=int_dia_sch40_3inch,
                        fitt_equiv=t_fitt_sch40_3inch,
                    )

                    og_flowrate_end += flow_rate_end
    
                    p_drop = pressure_drop(
                        q=og_flowrate_end, l_eqv=length, d=cr_line_int_dia
                    )
                    pressure_add = all_pressure[len(all_pressure) - 2] + p_drop
                    inverted_pressure_add.append(pressure_add)

                    all_flowrate.append(og_flowrate_end)

                    og1_flowrate_end = all_flowrate[len(all_flowrate) - 3]

                    og2_flow_rate_end_corr = (
                        og1_flowrate_end
                        * math.sqrt(
                            inverted_pressure_add[len(inverted_pressure_add) - 1]
                        )
                        / math.sqrt(main_pressure[len(main_pressure) - 2])
                    )
                    v += 6
                    inverted_flowrate.append(og2_flow_rate_end_corr)
                    og_flowrate_end += og2_flow_rate_end_corr
                elif n == no_of_sprinkler_on_branch_lines - 1:
                    length = equiv_length_t(
                        l=DIST_OF_SPRINKLER_ON_BRANCH_LINES / 2,
                        d_1=br2_line_int_dia,
                        d_2=int_dia_sch40_2inch,
                        fitt_equiv=t_fitt_sch40_2inch,
                    )
                    p_drop = pressure_drop(
                        q=og2_flow_rate_end_corr, l_eqv=length, d=br2_line_int_dia
                    )
                    pressure_add = inverted_pressure_add[0] - p_drop
                    inverted_pressure_add.append(pressure_add)
                    discharge = discharge_rate(p=pressure_add)
                    inverted_discharge_add.append(discharge)
                    og2_flow_rate_end_corr -= discharge
                    inverted_flowrate.append(og2_flow_rate_end_corr)
                    # owrinverted_flate.append(og2_flow_rate_end_corr)

                elif n <= no_of_sprinkler_on_branch_lines - 2:
                    if n == no_of_sprinkler_on_branch_lines - 2:
                        p_drop = pressure_drop(
                            q=og2_flow_rate_end_corr,
                            l_eqv=DIST_OF_SPRINKLER_ON_BRANCH_LINES,
                            d=br2_line_int_dia,
                        )
                        pressure_add = (
                            inverted_pressure_add[len(inverted_pressure_add) - 1]
                            - p_drop
                        )
                        inverted_pressure_add.append(pressure_add)
                        discharge = discharge_rate(p=pressure_add)
                        inverted_discharge_add.append(discharge)
                        og2_flow_rate_end_corr -= discharge
                        inverted_flowrate.append(og2_flow_rate_end_corr)
                   
                    else:
                        
                        p_drop = pressure_drop(
                            q=og2_flow_rate_end_corr,
                            l_eqv=DIST_OF_SPRINKLER_ON_BRANCH_LINES,
                            d=br_line_int_dia,
                        )
                        pressure_add = (
                            inverted_pressure_add[len(inverted_pressure_add) - 1]
                            - p_drop
                        )
                        inverted_pressure_add.append(pressure_add)
                       
                        if n!=1:
                            discharge = discharge_rate(p=pressure_add)
                            inverted_discharge_add.append(discharge)
                            og2_flow_rate_end_corr -= discharge
                            
                            inverted_flowrate.append(og2_flow_rate_end_corr)
                        elif n==1:
                            inverted_discharge_add.append(discharge)
                            
                    
                        
                        
                        if n == 1:
                            length = equiv_length_t(
                                l=DIST_OF_SPRINKLER_ON_BRANCH_LINES / 2,
                                d_1=br_line_int_dia,
                                d_2=int_dia_sch40_1_5inch,
                                fitt_equiv=t_fitt_sch40_1_5inch,
                            )
                            p_drop = pressure_drop(
                                q=node1_discharge_rate, l_eqv=length, d=br_line_int_dia
                            )
                        
                            all_pressure.extend(reversed(inverted_pressure_add))
                            all_flowrate.extend(reversed(inverted_flowrate))
                            discharge = (
                                node1_discharge_rate
                                * math.sqrt(all_pressure[len(all_pressure) - 1])
                                / math.sqrt(node1_pressure + p_drop)
                            )
                            all_discharge_rate.extend(reversed(inverted_discharge_add))
                            all_discharge_rate.append(discharge)
                            all_flowrate.append(discharge)
                            p_drop = pressure_drop(
                                q=discharge, l_eqv=length, d=br_line_int_dia
                            )
                            all_pressure.append(
                                all_pressure[len(all_pressure) - 1] - p_drop
                            )
                            inverted_discharge_add = []
                            inverted_pressure_add = []
                            inverted_flowrate = []
                            og_flowrate_end += all_discharge_rate[
                                len(all_discharge_rate) - 1]
                            
                            if m == no_of_branch_lines:
                                all_flowrate.append(og_flowrate_end)
                            
                            flow_rate_end = og_flowrate_end

   
    p_drop_elev = elevation * 0.433
    system_pressure = all_pressure[len(all_pressure) - 2] + p_drop_elev
    fire_pump_HP = horsepower_calc(
        all_flowrate[len(all_flowrate) - 1], all_pressure[len(all_pressure) - 2]
    )
    fire_pump_cap = suitable_pump(og_flowrate_end)
    supply_time = 60

    """preparation of data for the table"""
    row_data = [
        [
            "total number of sprinklers",
            no_of_sprinkler_on_branch_lines * no_of_branch_lines,
        ],
        ["total number of branch lines", no_of_branch_lines],
        ["total number of sprinklers on branch lines", no_of_sprinkler_on_branch_lines],
        [
            "total number of sprinklers in design area",
            min_sprinkler_on_branch * sprinkler_across_branch + add_node,
        ],
        ["flow density(gpm/ft²)", density],
        ["hose allowance (gpm)", hose_stream_demand],
        ["system flow(gpm)", all_flowrate[len(all_flowrate) - 1]],
        ["system pressure(psi)", system_pressure],
    ]

    df1 = pd.DataFrame(row_data)

    row_data2 = [
        ["Power of fire pump(HP)", fire_pump_HP],
        ["Capacity of fire pump(gpm)", fire_pump_cap],
        ["Power of Jockey pump(HP)", 0.1 * fire_pump_HP],
        [" Capacity of Jockey pump(gpm)", suitable_pump(node1_discharge_rate)],
        ["Minimum Tank size(m³)", fire_pump_cap * supply_time / 264.2],
        ["water supply duration(min)", supply_time],
    ]

    df2 = pd.DataFrame(row_data2)

    app = [2, 3, 4, 6, 7]
    count = 1
    index = range(1, len(all_pressure) + 1)
    heading = [
        "start node",
        "end node",
        "start node pressure",
        "end node pressure",
        "flow",
        "pipe diameter",
        "pipe length",
        "head loss",
    ]
    headings_data = []
    for n in index:
        if n <= no_of_sprinkler_on_branch_lines + 1:
            if n < no_of_sprinkler_on_branch_lines - 2:
                entry = [
                    n,
                    n + 1,
                    all_pressure[n - 1],
                    all_pressure[n],
                    all_flowrate[n - 1],
                    br_line_int_dia,
                    DIST_OF_SPRINKLER_ON_BRANCH_LINES,
                ]
                entry.append(entry[3] - entry[2])
                for i, j in enumerate(app):
                    entry[j] = format(entry[j], ".3f")
                headings_data.append(entry)

            elif n == no_of_sprinkler_on_branch_lines - 2:
                entry = [
                    n,
                    n + 1,
                    all_pressure[n - 1],
                    all_pressure[n],
                    all_flowrate[n - 1],
                    br2_line_int_dia,
                    DIST_OF_SPRINKLER_ON_BRANCH_LINES,
                ]
                entry.append(entry[3] - entry[2])
                for i, j in enumerate(app):
                    entry[j] = format(entry[j], ".3f")
                headings_data.append(entry)

            elif n == no_of_sprinkler_on_branch_lines - 1:
                length = equiv_length_t(
                    l=DIST_OF_SPRINKLER_ON_BRANCH_LINES / 2,
                    d_1=br2_line_int_dia,
                    d_2=int_dia_sch40_2inch,
                    fitt_equiv=t_fitt_sch40_2inch,
                )

                entry = [
                    n,
                    n + 1,
                    all_pressure[n - 1],
                    all_pressure[n],
                    all_flowrate[n - 1],
                    br2_line_int_dia,
                    length,
                ]
                entry.append(entry[3] - entry[2])
                for i, j in enumerate(app):
                    entry[j] = format(entry[j], ".3f")
                headings_data.append(entry)

            elif n == no_of_sprinkler_on_branch_lines:
                length = equiv_length_t(
                    l=dist_between_branch_lines,
                    d_1=cr_line_int_dia,
                    d_2=int_dia_sch40_3inch,
                    fitt_equiv=t_fitt_sch40_3inch,
                )
                entry = [
                    n,
                    n + (no_of_sprinkler_on_branch_lines + 1),
                    all_pressure[n - 1],
                    all_pressure[n + no_of_sprinkler_on_branch_lines],
                    all_flowrate[n],
                    br_line_int_dia,
                    length,
                ]
                entry.append(entry[3] - entry[2])
                for i, j in enumerate(app):
                    entry[j] = format(entry[j], ".3f")
                headings_data.append(entry)
            elif n == no_of_sprinkler_on_branch_lines + 1:
                length = equiv_length_t(
                    l=DIST_OF_SPRINKLER_ON_BRANCH_LINES / 2,
                    d_1=br_line_int_dia,
                    d_2=int_dia_sch40_1_5inch,
                    fitt_equiv=t_fitt_sch40_1_5inch,
                )

                entry = [
                    n,
                    n - 1,
                    all_pressure[n - 1],
                    all_pressure[n - 2],
                    all_flowrate[n - 2],
                    br_line_int_dia,
                    length,
                ]
                entry.append(entry[3] - entry[2])
                for i, j in enumerate(app):
                    entry[j] = format(entry[j], ".3f")
                headings_data.append(entry)

                count += 1

        elif n > no_of_sprinkler_on_branch_lines + 1:
            p = n % (no_of_sprinkler_on_branch_lines + 1)
            if p == 0:
                p = no_of_sprinkler_on_branch_lines + 1
            if p <= no_of_sprinkler_on_branch_lines + 1:
                if p < no_of_sprinkler_on_branch_lines - 2:
                    entry = [
                        n,
                        n + 1,
                        all_pressure[n - 1],
                        all_pressure[n],
                        all_flowrate[n - 1],
                        br_line_int_dia,
                        DIST_OF_SPRINKLER_ON_BRANCH_LINES,
                    ]
                    entry.append(entry[3] - entry[2])
                    for i, j in enumerate(app):
                        entry[j] = format(entry[j], ".3f")
                    headings_data.append(entry)

                elif p == no_of_sprinkler_on_branch_lines - 2:
                    entry = [
                        n,
                        n + 1,
                        all_pressure[n - 1],
                        all_pressure[n],
                        all_flowrate[n - 1],
                        br2_line_int_dia,
                        DIST_OF_SPRINKLER_ON_BRANCH_LINES,
                    ]
                    entry.append(entry[3] - entry[2])
                    for i, j in enumerate(app):
                        entry[j] = format(entry[j], ".3f")
                    headings_data.append(entry)

                elif p == no_of_sprinkler_on_branch_lines - 1:
                    length = equiv_length_t(
                        l=DIST_OF_SPRINKLER_ON_BRANCH_LINES / 2,
                        d_1=br2_line_int_dia,
                        d_2=int_dia_sch40_2inch,
                        fitt_equiv=t_fitt_sch40_2inch,
                    )

                    entry = [
                        n,
                        n + 1,
                        all_pressure[n - 1],
                        all_pressure[n],
                        all_flowrate[n - 1],
                        br2_line_int_dia,
                        length,
                    ]
                    entry.append(entry[3] - entry[2])
                    for i, j in enumerate(app):
                        entry[j] = format(entry[j], ".3f")
                    headings_data.append(entry)

                elif p == no_of_sprinkler_on_branch_lines:
                    if count < no_of_branch_lines:
                        length = equiv_length_t(
                            l=dist_between_branch_lines,
                            d_1=cr_line_int_dia,
                            d_2=int_dia_sch40_3inch,
                            fitt_equiv=t_fitt_sch40_3inch,
                        )

                        entry = [
                            n,
                            n + (no_of_sprinkler_on_branch_lines + 1),
                            all_pressure[n - 1],
                            all_pressure[n + no_of_sprinkler_on_branch_lines],
                            all_flowrate[n],
                            br_line_int_dia,
                            length,
                        ]
                        entry.append(entry[3] - entry[2])
                        for i, j in enumerate(app):
                            entry[j] = format(entry[j], ".3f")
                        headings_data.append(entry)
                elif p == no_of_sprinkler_on_branch_lines + 1:
                    length = equiv_length_t(
                        l=DIST_OF_SPRINKLER_ON_BRANCH_LINES / 2,
                        d_1=br_line_int_dia,
                        d_2=int_dia_sch40_1_5inch,
                        fitt_equiv=t_fitt_sch40_1_5inch,
                    )

                    entry = [
                        n,
                        n - 1,
                        all_pressure[n - 1],
                        all_pressure[n - 2],
                        all_flowrate[n - 2],
                        br_line_int_dia,
                        length,
                    ]
                    entry.append(entry[3] - entry[2])
                    for i, j in enumerate(app):
                        entry[j] = format(entry[j], ".3f")
                    headings_data.append(entry)
                    count += 1



    df3 = pd.DataFrame(
        headings_data, columns=heading, index=range(1, len(headings_data) + 1)
    )

    headings_data2 = []
    index2 = range(1, len(all_pressure) + 1)

    heading2 = [
        "node point",
        "pressure",
        "discharge rate",
    ]
    app = [1, 2]
    for n in index:
        """loop to format the data to 3 decimal places"""
        entry = [n, all_pressure[n - 1], all_discharge_rate[n - 1]]
        for i, j in enumerate(app):
            entry[j] = format(entry[j], ".3f")
        headings_data2.append(entry)

    df4 = pd.DataFrame(headings_data2, columns=heading2, index=index2)

    wb = openpyxl.Workbook()
    wb.create_sheet("General data")
    wb.create_sheet("Pump&Tank data")
    wb.create_sheet("pipe data")
    wb.create_sheet("node data")
    wb.save("pipe_data.xlsx")

    with pd.ExcelWriter("pipe_data.xlsx") as writer:
        df1.to_excel(writer, sheet_name="General data", index=0, header= False)
        df2.to_excel(writer, sheet_name="Pump&Tank data", index=0, header= False)
        df3.to_excel(writer, sheet_name="pipe data" )
        df4.to_excel(writer, sheet_name="node data")
